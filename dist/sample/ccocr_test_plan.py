# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'テスト計画'

# --- スタイル定義 ---
hdr_fill   = PatternFill('solid', fgColor='C0392B')   # 赤
cat_fill   = PatternFill('solid', fgColor='FADBD8')   # 薄赤
item_fill  = PatternFill('solid', fgColor='FDFEFE')   # 白
hdr_font   = Font(bold=True, color='FFFFFF', size=11)
cat_font   = Font(bold=True, size=10)
item_font  = Font(size=10)
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border

def cat(ws, row, label):
    c = ws.cell(row=row, column=1, value=label)
    c.fill = cat_fill; c.font = cat_font
    c.alignment = Alignment(vertical='center')
    c.border = border
    for col in range(2, 7):
        wc = ws.cell(row=row, column=col)
        wc.fill = cat_fill; wc.border = border

def item(ws, row, no, test, check, note=''):
    data = [no, test, check, '', note]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = item_fill; c.font = item_font
        c.border = border
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if col == 4:  # チェック列
            c.alignment = Alignment(horizontal='center', vertical='center')
        if col == 3:  # 確認内容
            c.alignment = Alignment(vertical='center', wrap_text=True)

# --- ヘッダ行 ---
ws.row_dimensions[1].height = 30
hdr(ws, 1, 1, 'No.')
hdr(ws, 1, 2, 'テスト項目')
hdr(ws, 1, 3, '確認内容・手順')
hdr(ws, 1, 4, '結果\n(✓/✗)')
hdr(ws, 1, 5, 'メモ')

# --- テストデータ ---
rows = [
    # (type, args...)
    ('cat', '1. 起動・インストール'),
    ('item', 1,  '初回起動（フレッシュ）',
     'exe + MinGit フォルダを同じ場所に置いて exe 実行\n'
     '→ AppData\\ChuanlaiApps\\ccocr にコードがcloneされる\n'
     '→「インストールが完了しました」メッセージ表示'),
    ('item', 2,  '2回目起動',
     '同じ exe を再実行\n'
     '→ git pull が走る（コンソールに表示される）\n'
     '→ 設定Excelダイアログが開く'),
    ('item', 3,  '設定Excel記憶',
     '設定Excelを選択して処理完了後、再度 exe 実行\n'
     '→ 設定Excelダイアログが出ずに直接起動する'),

    ('cat', '2. Python・モジュール'),
    ('item', 4,  'Python検出',
     'コンソールに「Python: Python 3.x.x」が表示される'),
    ('item', 5,  'モジュール確認',
     'コンソールに「module ready: xxx」が各モジュール分表示される'),

    ('cat', '3. 基本パイプライン（frmモード）'),
    ('item', 6,  '設定Excel読み込み',
     '設定Excelを選択後、エラーなく次の処理に進む'),
    ('item', 7,  'PDF→PNG変換',
     'pngPRE/ および pngUP/ にPNGが生成される'),
    ('item', 8,  'Azure API呼び出し',
     'jsnRAW/ にJSONファイルが保存される\n'
     '（CV: xxx_cv.json / DI: xxx_di.json）'),
    ('item', 9,  'データ抽出→SQLite',
     'dbf/ に .db ファイルが生成される'),
    ('item', 10, 'Excel出力（画像付き）',
     'output/ に jobid.xlsm が生成される'),
    ('item', 11, 'WEB確認画面',
     'ブラウザが開き確認画面が表示される\n'
     '→ 確認ボタン押下後 jobid_WEB.xlsm が生成される'),

    ('cat', '4. 各種組み合わせ'),
    ('item', 12, 'DI / cnv (2d lv0)',
     '文字列抽出 ✓ / spic画像 ✓'),
    ('item', 13, 'DI / ncv (nd lvn)',
     '文字列抽出 ✓ / spic画像 ✓'),
    ('item', 14, 'CV / cnv (2d lv0)',
     '文字列抽出 ✓ / spic画像 ✓'),
    ('item', 15, 'CV / ncv (nd lvn)',
     '文字列抽出 ✓ / spic画像 ✓'),
    ('item', 16, '回転PNG (rot.png)',
     '-104°回転画像を入力\n'
     '→ 回転補正されて正常にデータ抽出される'),
    ('item', 17, 'マルチページ + RG',
     '3ページ+空白ページのPDFを入力\n'
     '→ 繰り返しグループが正常に抽出される'),

    ('cat', '5. git pull 自動更新'),
    ('item', 18, 'コード更新の反映',
     'Mac側でコードを1行変更してgit push\n'
     '→ Win10で exe 起動\n'
     '→ git pull で変更が反映されていること確認'),
]

row = 2
for r in rows:
    if r[0] == 'cat':
        ws.row_dimensions[row].height = 20
        cat(ws, row, r[1])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    else:
        ws.row_dimensions[row].height = 55
        item(ws, row, r[1], r[2], r[3])
    row += 1

# --- 列幅 ---
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 48
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 24

# --- ウィンドウ枠固定 ---
ws.freeze_panes = 'A2'

output = '/Users/cy/Desktop/CLAUDEs/ccocr/dist/sample/ccocr_test_plan.xlsx'
wb.save(output)
print(f'saved: {output}')
